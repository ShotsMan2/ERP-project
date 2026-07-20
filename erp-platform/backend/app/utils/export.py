from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any


def export_to_csv(data: list[dict[str, Any]], columns: list[str] | None = None) -> str:
    if not data:
        return ""
    if columns is None:
        columns = list(data[0].keys())
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in data:
        writer.writerow([str(row.get(col, "")) for col in columns])
    return output.getvalue()


def export_to_csv_file(data: list[dict[str, Any]], filepath: str | Path, columns: list[str] | None = None) -> str:
    csv_content = export_to_csv(data, columns)
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(csv_content, encoding="utf-8-sig")
    return str(path)


def export_to_excel(data: list[dict[str, Any]], columns: list[str] | None = None) -> bytes:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        wb.create_sheet()
        ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    if not data:
        wb.save(io.BytesIO())
        return b""

    if columns is None:
        columns = list(data[0].keys())

    header_font = openpyxl.styles.Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(col_name, "")))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_excel_file(data: list[dict[str, Any]], filepath: str | Path, columns: list[str] | None = None) -> str:
    content = export_to_excel(data, columns)
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return str(path)
